import os
import io
import datetime
import requests
from flask import Flask, render_template, request, Response
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from itertools import groupby

# ======================
# Konfiguration
# ======================
NETBOX_URL = os.getenv("NETBOX_URL", "")
NETBOX_TOKEN = os.getenv("NETBOX_TOKEN", "")  # leer bei Demo
EXPORT_DIR = "exports"

HEADERS = {}
if NETBOX_TOKEN:
    HEADERS["Authorization"] = f"Token {NETBOX_TOKEN}"

os.makedirs(EXPORT_DIR, exist_ok=True)

app = Flask(__name__)


# ======================
# Hilfsfunktionen
# ======================
def netbox_get(endpoint, params=None):
    url = f"{NETBOX_URL}/api/{endpoint}"
    response = requests.get(url, headers=HEADERS, params=params)
    response.raise_for_status()
    return response.json()


def get_all_results(endpoint, params=None):
    """Behandelt Pagination von NetBox"""
    results = []
    while True:
        data = netbox_get(endpoint, params)
        results.extend(data["results"])
        if not data["next"]:
            break
        endpoint = data["next"].replace(f"{NETBOX_URL}/api/", "")
        params = None
    return results


# ======================
# Routes
# ======================
@app.route("/bom-export", methods=["GET"])
def index():
    sites = get_all_results("dcim/sites/")
    sites = sorted(sites, key=lambda s: s["name"])
    return render_template("index.html", sites=sites)


@app.route("/bom-export/export", methods=["POST"])
def export():
    site_slug = request.form.get("site_slug")
    site_name = request.form.get("site_name")

    devices = get_all_results(
        "dcim/devices/",
        params={"site": site_slug, "limit": 1000}
    )

    rows = []
    for d in devices:
        rows.append({
            "Location": d["location"]["name"] if d.get("location") else "",
            "Rack": d["rack"]["name"] if d.get("rack") else "",
            "Device Name": d.get("name", ""),
            "Manufacturer": (
                d["device_type"]["manufacturer"]["name"]
                if d.get("device_type") and d["device_type"].get("manufacturer")
                else ""
            ),
            "Device Type": (
                d["device_type"]["display"]
                if d.get("device_type")
                else ""
            ),
            "Serial Number": d.get("serial", ""),
            "Role": (
                d["role"]["name"]
                if d.get("role")
                else ""
            )
        })

    # Sort for grouping
    rows.sort(key=lambda x: (x["Location"] or "", x["Rack"] or "", x["Device Name"] or ""))

    # Excel Generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Define styles
    site_header_font = Font(bold=True, size=16)
    location_header_font = Font(bold=True, size=14)
    rack_header_font = Font(bold=True, size=12)
    header_font = Font(bold=True)

    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    current_row = 1

    # 1. Site Header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(row=current_row, column=1, value=f"Site: {site_name}")
    cell.font = site_header_font
    cell.alignment = center_alignment
    current_row += 1

    def get_location(d): return d["Location"] or ""
    def get_rack(d): return d["Rack"] or ""

    for location, loc_devices in groupby(rows, key=get_location):
        loc_devices = list(loc_devices)

        # 2. Location Header
        display_location = location if location else "No Location"

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        cell = ws.cell(row=current_row, column=1, value=f"Location: {display_location}")
        cell.font = location_header_font
        cell.alignment = left_alignment
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        current_row += 1

        # Group by Rack
        for rack, rack_devices in groupby(loc_devices, key=get_rack):
            rack_devices = list(rack_devices)

            # 3. Rack Header (only if rack exists)
            if rack:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                cell = ws.cell(row=current_row, column=1, value=f"Rack: {rack}")
                cell.font = rack_header_font
                cell.alignment = left_alignment
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                current_row += 1

            # 4. Column Headers
            headers = ["Device Name", "Manufacturer", "Device Type", "Serial Number", "Role"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.border = border
            current_row += 1

            # 5. Device Rows
            for d in rack_devices:
                ws.cell(row=current_row, column=1, value=d["Device Name"]).border = border
                ws.cell(row=current_row, column=2, value=d["Manufacturer"]).border = border
                ws.cell(row=current_row, column=3, value=d["Device Type"]).border = border
                ws.cell(row=current_row, column=4, value=d["Serial Number"]).border = border
                ws.cell(row=current_row, column=5, value=d["Role"]).border = border
                current_row += 1

            # Space between groups
            current_row += 1

    # Auto-adjust column widths
    for i, col in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Dateiname
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    safe_site = site_name.replace(" ", "_")
    filename = f"{safe_site}_BOM_{date_str}.xlsx"

    # Excel erzeugen
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()

    # Archivieren
    with open(os.path.join(EXPORT_DIR, filename), "wb") as f:
        f.write(excel_bytes)

    # Download
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        excel_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# ======================
# Start
# ======================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
